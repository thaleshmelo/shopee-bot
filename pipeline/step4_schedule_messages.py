# pipeline/step4_schedule_messages.py
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


TOTAL_MESSAGES = int(os.getenv("TOTAL_MESSAGES", "15"))
TIME_BLOCKS = [("09:00", "14:00", 5), ("14:00", "20:00", 5), ("20:00", "23:30", 5)]

# Se SHOPEE_MESSAGES_FILE estiver definido, ele tem prioridade e o Step4 usa esse arquivo diretamente.
MESSAGES_FILE = os.getenv("SHOPEE_MESSAGES_FILE", "").strip()

# Default: sempre usar o arquivo final com links
DEFAULT_PICKS = os.path.join("outputs", "picks_refinados_com_links.csv")
PICKS_FILE = os.getenv("SHOPEE_PICKS_FILE", DEFAULT_PICKS).strip()

MIN_SHORTLINK_COVERAGE = float(os.getenv("STEP4_MIN_SHORTLINK_COVERAGE", "0.90"))


def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _clean_str(s: pd.Series) -> pd.Series:
    s = s.astype(str).fillna("").str.strip()
    s = s.replace({"nan": "", "None": "", "NULL": "", "null": ""})
    return s


def parse_time(t: str) -> datetime:
    today = datetime.now().date()
    h, m = map(int, t.split(":"))
    return datetime.combine(today, datetime.min.time()).replace(hour=h, minute=m)


def distribute_times():
    times = []
    for start, end, qty in TIME_BLOCKS:
        start_dt = parse_time(start)
        end_dt = parse_time(end)
        delta = (end_dt - start_dt) / qty
        for i in range(qty):
            times.append(start_dt + delta * i)
    return times[:TOTAL_MESSAGES]


def save_xlsx_pretty(df: pd.DataFrame, path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Agenda"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # header style
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # column widths (simples)
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, val in enumerate(row, start=1):
            widths[i] = max(widths.get(i, 10), len(str(val)) if val is not None else 0)
    for i, w in widths.items():
        ws.column_dimensions[chr(64 + i)].width = min(max(w + 2, 12), 60)

    # table style
    tab = Table(displayName="AgendaTable", ref=f"A1:{chr(64+ws.max_column)}{ws.max_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save(path)


def read_input_dataframe() -> pd.DataFrame:
    # 1) Se for arquivo de mensagens pronto (CSV/XLSX), usa ele.
    if MESSAGES_FILE:
        p = Path(MESSAGES_FILE)
        if not p.exists():
            raise SystemExit(f"SHOPEE_MESSAGES_FILE não encontrado: {p}")
        if p.suffix.lower() == ".csv":
            return normalize_cols(pd.read_csv(p, low_memory=False))
        if p.suffix.lower() == ".xlsx":
            return normalize_cols(pd.read_excel(p))
        raise SystemExit("SHOPEE_MESSAGES_FILE deve ser .csv ou .xlsx")

    # 2) Picks com links
    p = Path(PICKS_FILE)
    if not p.exists():
        raise SystemExit(f"Arquivo de picks não encontrado: {p}")

    print(f"LENDO PICKS_FILE: {PICKS_FILE}")
    return normalize_cols(pd.read_csv(p, low_memory=False))


def main() -> None:
    df = read_input_dataframe()

    # Modo mensagens prontas
    if MESSAGES_FILE:
        # precisa ter ao menos coluna "message"
        if "message" not in df.columns:
            raise SystemExit(f"SHOPEE_MESSAGES_FILE precisa ter coluna 'message'. Colunas: {list(df.columns)}")
        df["message"] = _clean_str(df["message"])
        df = df[df["message"].str.len() > 0].copy()
        times = distribute_times()

        rows = []
        for i, t in enumerate(times):
            msg = df.iloc[i % len(df)]["message"]
            rows.append({"time": t.strftime("%H:%M"), "message": msg})

        out = pd.DataFrame(rows)
    else:
        # Modo a partir de picks
        required_any = ["product_link", "product_short_link"]
        if not any(c in df.columns for c in required_any):
            raise SystemExit(f"Arquivo de picks não tem colunas de link. Colunas: {list(df.columns)}")

        # Normaliza colunas esperadas
        if "title" not in df.columns:
            # tenta nomes alternativos
            for alt in ["nome_curto", "productName", "offerName", "name"]:
                if alt in df.columns:
                    df["title"] = df[alt]
                    break
        if "sale_price" not in df.columns:
            for alt in ["preco_atual", "price", "salePrice"]:
                if alt in df.columns:
                    df["sale_price"] = df[alt]
                    break

        if "product_link" not in df.columns:
            # tenta achar algum link
            for alt in ["link_afiliado", "productLink", "offerLink", "originalLink", "url", "link"]:
                if alt in df.columns:
                    df["product_link"] = df[alt]
                    break

        if "product_short_link" not in df.columns:
            df["product_short_link"] = ""

        df["product_link"] = _clean_str(df["product_link"])
        df["product_short_link"] = _clean_str(df["product_short_link"])

        # ✅ FIX: fallback no Step4 também (antes de medir cobertura)
        df.loc[df["product_short_link"].str.len() == 0, "product_short_link"] = df["product_link"]

        total = len(df)
        filled = int(df["product_short_link"].str.len().gt(0).sum())
        coverage = (filled / total) if total else 0.0

        if coverage < MIN_SHORTLINK_COVERAGE:
            raise SystemExit(
                f"Short links insuficientes: {coverage:.0%} preenchidos (mínimo {int(MIN_SHORTLINK_COVERAGE*100)}%). "
                f"Verifique o Step3 / arquivo {PICKS_FILE}."
            )

        # Gera agenda
        times = distribute_times()
        rows = []

        for i, t in enumerate(times):
            r = df.iloc[i % len(df)]
            title = str(r.get("title", "")).strip()
            price = r.get("sale_price", "")
            link = str(r.get("product_short_link", "")).strip()

            # mensagem simples e consistente
            msg = (
                "Se liga nessa oferta 👇\n"
                f"✅ {title}\n"
                f"💰 R$ {price}\n"
                "👉 Pega antes que mude o preço\n"
                f"🔗 {link}"
            )
            rows.append({"time": t.strftime("%H:%M"), "title": title, "price": price, "link": link, "message": msg})

        out = pd.DataFrame(rows)

    base = f"agenda_envios_{datetime.now().strftime('%Y-%m-%d')}"
    csv_name = f"{base}.csv"
    xlsx_name = f"{base}.xlsx"

    out.to_csv(csv_name, index=False, encoding="utf-8-sig")
    save_xlsx_pretty(out, xlsx_name)

    print(f"\n📁 Salvo: {csv_name}")
    print(f"📁 Salvo: {xlsx_name} (formatado)")


if __name__ == "__main__":
    main()
